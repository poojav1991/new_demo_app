import openpyxl


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    # sheet = workbook.active
    return sheet.max_row


def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    # sheet = workbook.active
    return sheet.max_column


def get_read_data(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    # sheet = workbook.active
    return sheet.cell(row=rownum, column=colnum).value


def get_write_data(file, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    # sheet = workbook.active
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)
